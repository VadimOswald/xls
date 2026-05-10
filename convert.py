import xlrd
import openpyxl
from collections import OrderedDict

PRODUCT_ORDER = ['минтай', 'хек', 'красная рыба', 'котлеты рыбные', 'рыбные консервы']

# Column mappings for each year sheet (0-indexed)
# Based on actual file structure:
# Natural (kg): D=3, E=4, F=5, G=6, H=7  (минтай, хек, красная рыба, котлеты, консервы)
# Cost (rub): J=9, K=10, L=11, M=12, N=13
# Count (person): P=15, Q=16, R=17, S=18, T=19
# Consumption (kg/person): V=21, W=22, X=23, Y=24, Z=25

NAT_COLS = [3, 4, 5, 6, 7]      # D, E, F, G, H
COST_COLS = [9, 10, 11, 12, 13]  # J, K, L, M, N
COUNT_COLS = [15, 16, 17, 18, 19]  # P, Q, R, S, T
CONS_COLS = [21, 22, 23, 24, 25]  # V, W, X, Y, Z

YEARS = ['2020', '2021', '2022', '2023', '2024', '2025']
SHEET_NAMES = ['Таблица2020', 'Таблица2021', 'Таблица2022', 'Таблица2023', 'Таблица2024', 'Таблица2025']

def normalize_name(name):
    """Normalize institution name by stripping whitespace"""
    if isinstance(name, str):
        return ' '.join(name.split())
    return str(name)

def get_cell_value(sh, r, c):
    """Get cell value, return 0 for empty numeric cells"""
    try:
        val = sh.cell_value(r, c)
        if val is None or val == '':
            return 0
        return val
    except:
        return 0

# Load source data
wb = xlrd.open_workbook('отсюда.xls')

# Data structure: {normalized_inst_name: {product: {year: {nat, cost, count, cons}}}}
data = OrderedDict()
inst_order = []  # To preserve order from 2020 sheet

for year_idx, sheet_name in enumerate(SHEET_NAMES):
    sh = wb.sheet_by_name(sheet_name)
    year = YEARS[year_idx]
    
    # Find data rows (skip header rows 0-5)
    for r in range(6, sh.nrows):
        inst_name = sh.cell_value(r, 1)
        if not inst_name or not isinstance(inst_name, str) or len(inst_name.strip()) < 3:
            continue
        
        # Skip total/summary rows
        if 'итого' in str(inst_name).lower() or 'всего' in str(inst_name).lower():
            continue
            
        norm_inst = normalize_name(inst_name)
        
        # Track order from first year only
        if year_idx == 0 and norm_inst not in inst_order:
            inst_order.append(norm_inst)
            data[norm_inst] = {p: {y: {'nat': 0, 'cost': 0, 'count': 0, 'cons': 0} for y in YEARS} for p in PRODUCT_ORDER}
        
        if norm_inst not in data:
            data[norm_inst] = {p: {y: {'nat': 0, 'cost': 0, 'count': 0, 'cons': 0} for y in YEARS} for p in PRODUCT_ORDER}
        
        # Read data for each product
        for prod_idx, prod in enumerate(PRODUCT_ORDER):
            nat_val = get_cell_value(sh, r, NAT_COLS[prod_idx])
            cost_val = get_cell_value(sh, r, COST_COLS[prod_idx])
            count_val = get_cell_value(sh, r, COUNT_COLS[prod_idx])
            cons_val = get_cell_value(sh, r, CONS_COLS[prod_idx])
            
            # Sum values if multiple rows exist
            data[norm_inst][prod][year]['nat'] += nat_val
            data[norm_inst][prod][year]['cost'] += cost_val
            data[norm_inst][prod][year]['count'] += count_val
            data[norm_inst][prod][year]['cons'] += cons_val

# Create output file
out_wb = openpyxl.Workbook()
sh_out = out_wb.active
sh_out.title = 'Лист1'

# Headers as specified
headers_row1 = [
    '№ п/п',
    'Наименование учреждения/\nорганизатора закупки',
    'вид продукции'
]

# Add year headers (4 columns per year: nat, cost, count, cons)
for year in YEARS:
    headers_row1.extend([
        f'{year} натуральное',
        f'{year} стоимость',
        f'{year} численность',
        f'{year} расчётное потребление'
    ])

# Write header row
for col, header in enumerate(headers_row1, 1):
    sh_out.cell(row=1, column=col, value=header)

# Write data rows
row_num = 2
for inst_name in inst_order:
    for prod in PRODUCT_ORDER:
        row_data = [row_num - 1, inst_name, prod]
        
        for year in YEARS:
            row_data.append(data[inst_name][prod][year]['nat'])
            row_data.append(data[inst_name][prod][year]['cost'])
            row_data.append(data[inst_name][prod][year]['count'])
            row_data.append(data[inst_name][prod][year]['cons'])
        
        for col, val in enumerate(row_data, 1):
            sh_out.cell(row=row_num, column=col, value=val)
        
        row_num += 1

# Save
out_wb.save('сюда.xlsx')
print(f'Done! Written {row_num - 2} data rows.')
print(f'Institutions: {len(inst_order)}')
print(f'Total rows: {len(inst_order) * len(PRODUCT_ORDER)}')
