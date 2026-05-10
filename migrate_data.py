#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Скрипт переноса данных из отсюда.xls в сюда.xlsx
"""

import pandas as pd
from openpyxl import Workbook

# Продукты в правильном порядке
PRODUCTS = ['минтай', 'хек', 'красная рыба', 'котлеты рыбные', 'рыбные консервы']

def parse_sheet(df, year):
    """Извлекаем данные из листа для конкретного года"""
    data = {}
    
    # Определяем смещение колонок в зависимости от года
    if year == 2020:
        org_col = None  # Нет отдельной колонки с организатором
        nat_start = 3
        cost_start = 9
        count_start = 15
        cons_start = 21
    else:
        org_col = 2
        nat_start = 4
        cost_start = 10
        count_start = 16
        cons_start = 22
    
    # Проходим по всем строкам с данными (начиная с строки 6)
    for i in range(6, len(df)):
        row = df.iloc[i]
        
        # Пропускаем строки без номера или итоговые строки
        if pd.isna(row.iloc[0]):
            continue
        
        try:
            row_num = str(row.iloc[0]).strip()
            if not row_num.isdigit():
                continue
            # Пропускаем итоговые строки
            if 'итог' in str(row.iloc[0]).lower() or 'всего' in str(row.iloc[1]).lower():
                continue
        except:
            continue
        
        institution = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ''
        if not institution:
            continue
            
        # Получаем организатора закупки (если есть)
        organizer = ''
        if org_col is not None and org_col < len(row):
            val = row.iloc[org_col]
            if pd.notna(val):
                organizer = str(val).strip()
        
        # Извлекаем данные для каждого продукта
        for prod_idx, product in enumerate(PRODUCTS):
            nat_idx = nat_start + prod_idx
            cost_idx = cost_start + prod_idx
            count_idx = count_start + prod_idx
            cons_idx = cons_start + prod_idx
            
            nat_val = row.iloc[nat_idx] if nat_idx < len(row) else 0
            cost_val = row.iloc[cost_idx] if cost_idx < len(row) else 0
            count_val = row.iloc[count_idx] if count_idx < len(row) else 0
            cons_val = row.iloc[cons_idx] if cons_idx < len(row) else 0
            
            # Ключ для группировки: учреждение + продукт
            key = (institution, product)
            
            if key not in data:
                data[key] = {
                    'institution': institution,
                    'product': product,
                    'organizers': {},  # год -> организатор
                    'natural': {},     # год -> значение
                    'cost': {},        # год -> значение
                    'count': {},       # год -> значение
                    'consumption': {}  # год -> значение
                }
            
            # Сохраняем данные (суммируем если уже есть)
            data[key]['organizers'][year] = organizer
            data[key]['natural'][year] = data[key]['natural'].get(year, 0) + (float(nat_val) if pd.notna(nat_val) else 0)
            data[key]['cost'][year] = data[key]['cost'].get(year, 0) + (float(cost_val) if pd.notna(cost_val) else 0)
            data[key]['count'][year] = data[key]['count'].get(year, 0) + (float(count_val) if pd.notna(count_val) else 0)
            data[key]['consumption'][year] = data[key]['consumption'].get(year, 0) + (float(cons_val) if pd.notna(cons_val) else 0)
    
    return data


def main():
    xls = pd.ExcelFile('отсюда.xls', engine='xlrd')
    
    # Собираем данные со всех листов
    all_data = {}
    years = [2020, 2021, 2022, 2023, 2024, 2025]
    sheet_names = ['Таблица2020', 'Таблица2021', 'Таблица2022', 'Таблица2023', 'Таблица2024', 'Таблица2025']
    
    # Сохраняем порядок учреждений как в Таблица2020
    institutions_order = []
    
    for year, sheet_name in zip(years, sheet_names):
        df = pd.read_excel(xls, sheet_name=sheet_name, header=None)
        sheet_data = parse_sheet(df, year)
        
        # Сохраняем порядок учреждений из первого листа (2020)
        if year == 2020:
            for key in sheet_data.keys():
                inst = key[0]
                if inst not in institutions_order:
                    institutions_order.append(inst)
        
        # Объединяем данные
        for key, val in sheet_data.items():
            if key not in all_data:
                all_data[key] = val
            else:
                # Добавляем данные за этот год
                all_data[key]['organizers'].update(val['organizers'])
                all_data[key]['natural'].update(val['natural'])
                all_data[key]['cost'].update(val['cost'])
                all_data[key]['count'].update(val['count'])
                all_data[key]['consumption'].update(val['consumption'])
    
    print(f"Всего уникальных комбинаций (учреждение × продукт): {len(all_data)}")
    print(f"Всего учреждений: {len(institutions_order)}")
    
    # Создаем итоговый DataFrame
    rows = []
    row_num = 0
    
    for inst in institutions_order:
        for product in PRODUCTS:
            key = (inst, product)
            if key not in all_data:
                continue
            
            val = all_data[key]
            row_num += 1
            
            # Формируем название учреждения с организаторами
            # Формат: "Наименование учреждения / Организатор2020 / Организатор2021 / ..."
            org_parts = []
            for y in years:
                org = val['organizers'].get(y, '')
                if not org:
                    org = '—'
                org_parts.append(org)
            
            # Для 2020 года организатора нет в исходных данных, ставим прочерк
            if org_parts[0] == '':
                org_parts[0] = '—'
            
            institution_with_orgs = f"{val['institution']} / {' / '.join(org_parts)}"
            
            row = {
                '№ п/п': row_num,
                'Наименование учреждения/организатора закупки': institution_with_orgs,
                'вид продукции': product,
            }
            
            # Добавляем данные по годам
            for y in years:
                nat = val['natural'].get(y, '')
                cost = val['cost'].get(y, '')
                count = val['count'].get(y, '')
                cons = val['consumption'].get(y, '')
                
                # Форматируем названия колонок
                row[f'{y} натуральное'] = nat if nat != 0 else nat
                row[f'{y} стоимость'] = cost if cost != 0 else cost
                row[f'{y} численность'] = count if count != 0 else count
                row[f'{y} расчётное потребление'] = cons if cons != 0 else cons
            
            rows.append(row)
    
    # Создаем DataFrame
    df_result = pd.DataFrame(rows)
    
    # Переупорядочиваем колонки согласно требованиям
    columns_order = ['№ п/п', 'Наименование учреждения/организатора закупки', 'вид продукции']
    for y in years:
        columns_order.extend([
            f'{y} натуральное',
            f'{y} стоимость',
            f'{y} численность',
            f'{y} расчётное потребление'
        ])
    
    df_result = df_result[columns_order]
    
    # Записываем в Excel
    wb = Workbook()
    ws = wb.active
    ws.title = 'Данные'
    
    # Заголовки
    headers = [
        '№ п/п',
        'Наименование учреждения/организатора закупки',
        'вид продукции',
        '2020 натуральное', '2020 стоимость', '2020 численность', '2020 расчётное потребление',
        '2021 натуральное', '2021 стоимость', '2021 численность', '2021 расчётное потребление',
        '2022 натуральное', '2022 стоимость', '2022 численность', '2022 расчётное потребление',
        '2023 натуральное', '2023 стоимость', '2023 численность', '2023 расчётное потребление',
        '2024 натуральное', '2024 стоимость', '2024 численность', '2024 расчётное потребление',
        '2025 натуральное', '2025 стоимость', '2025 численность', '2025 расчётное потребление',
    ]
    
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Данные
    for row_idx, row in enumerate(rows, 2):
        for col_idx, col_name in enumerate(columns_order, 1):
            val = row.get(col_name, '')
            ws.cell(row=row_idx, column=col_idx, value=val)
    
    # Сохраняем файл
    wb.save('сюда.xlsx')
    print(f"Файл 'сюда.xlsx' создан успешно!")
    print(f"Всего строк: {len(rows)}")
    
    # Проверка первых строк
    print("\nПервые 5 строк:")
    for i, row in enumerate(rows[:5]):
        print(f"{i+1}. {row['Наименование учреждения/организатора закупки'][:50]}... - {row['вид продукции']}")


if __name__ == '__main__':
    main()
